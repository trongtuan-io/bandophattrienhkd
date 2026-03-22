import io
from datetime import date
from flask import Blueprint, jsonify, session, send_file, render_template, request
from models import get_db
from controllers.auth import login_required

export_bp = Blueprint('export', __name__)


def _create_sheet(wb, ws, title, title_color, rows, headers, header_fill, header_font, thin_border):
    from openpyxl.styles import Font, Alignment
    from openpyxl.cell.cell import MergedCell

    col_count = len(headers)
    last_col_letter = chr(ord('A') + col_count - 1)

    ws.title = title
    ws.merge_cells(f'A1:{last_col_letter}1')
    ws['A1'] = f'{title.upper()} - {date.today().strftime("%d/%m/%Y")}'
    ws['A1'].font = Font(bold=True, size=14, color=title_color)
    ws['A1'].alignment = Alignment(horizontal='center')

    ws.merge_cells(f'A2:{last_col_letter}2')
    ws['A2'] = f'Tổng cộng: {len(rows)} hộ'
    ws['A2'].font = Font(italic=True, size=11)
    ws['A2'].alignment = Alignment(horizontal='center')

    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border

    for i, r in enumerate(rows, 1):
        row_num = i + 4
        values = [i, r['ten_chu_ho'], r['ten_cua_hang'], r['mst'],
                  r.get('cccd', ''), r.get('sdt', ''), r.get('so_tai_khoan', ''),
                  r['dia_chi'], r['phuong_xa'], r['quan_huyen'], r['ghi_chu'] or '']
        for col_idx, val in enumerate(values, 1):
            cell = ws.cell(row=row_num, column=col_idx, value=val)
            cell.border = thin_border
            if col_idx == 1:
                cell.alignment = Alignment(horizontal='center')

    for col in ws.columns:
        max_len = 0
        col_letter = None
        for cell in col:
            if isinstance(cell, MergedCell):
                continue
            if col_letter is None:
                col_letter = cell.column_letter
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        if col_letter:
            ws.column_dimensions[col_letter].width = min(max_len + 4, 40)


@export_bp.route('/api/export/excel')
@login_required
def export_excel():
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

    trang_thai_filter = request.args.get('trang_thai', 'all')

    db = get_db()
    with db.cursor() as cur:
        if trang_thai_filter == 'chua_dang_ky':
            cur.execute("""
                SELECT h.*, nv.ho_ten as nhan_vien_ten
                FROM ho_kinh_doanh h
                LEFT JOIN nhan_vien nv ON h.nhan_vien_id = nv.id
                WHERE h.trang_thai='chua_dang_ky'
                ORDER BY h.quan_huyen, h.phuong_xa
            """)
            rows_chua = cur.fetchall()
            rows_da = []
        elif trang_thai_filter == 'da_dang_ky':
            cur.execute("""
                SELECT h.*, nv.ho_ten as nhan_vien_ten
                FROM ho_kinh_doanh h
                LEFT JOIN nhan_vien nv ON h.nhan_vien_id = nv.id
                WHERE h.trang_thai='da_dang_ky'
                ORDER BY h.quan_huyen, h.phuong_xa
            """)
            rows_chua = []
            rows_da = cur.fetchall()
        else:
            cur.execute("""
                SELECT h.*, nv.ho_ten as nhan_vien_ten
                FROM ho_kinh_doanh h
                LEFT JOIN nhan_vien nv ON h.nhan_vien_id = nv.id
                WHERE h.trang_thai='chua_dang_ky'
                ORDER BY h.quan_huyen, h.phuong_xa
            """)
            rows_chua = cur.fetchall()
            cur.execute("""
                SELECT h.*, nv.ho_ten as nhan_vien_ten
                FROM ho_kinh_doanh h
                LEFT JOIN nhan_vien nv ON h.nhan_vien_id = nv.id
                WHERE h.trang_thai IN ('da_dang_ky', 'cho_duyet')
                ORDER BY h.quan_huyen, h.phuong_xa
            """)
            rows_da = cur.fetchall()

    wb = openpyxl.Workbook()
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="2C5282", end_color="2C5282", fill_type="solid")
    green_fill = PatternFill(start_color="27864e", end_color="27864e", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    headers = ['STT', 'Tên chủ hộ', 'Tên cửa hàng', 'MST', 'CCCD', 'SĐT', 'Số tài khoản',
               'Địa chỉ', 'Phường/Xã', 'Quận/Huyện', 'Ghi chú']

    # Sheet 1: Hộ chưa đăng ký
    ws1 = wb.active
    if rows_chua:
        _create_sheet(wb, ws1, 'Hộ chưa đăng ký', 'C0392B', rows_chua, headers, header_fill, header_font, thin_border)
    else:
        ws1.title = 'Hộ chưa đăng ký'
        ws1['A1'] = 'Không có hộ chưa đăng ký'

    # Sheet 2: Hộ đã đăng ký
    if rows_da:
        ws2 = wb.create_sheet()
        _create_sheet(wb, ws2, 'Hộ đã đăng ký', '27864e', rows_da, headers, green_fill, header_font, thin_border)

    filename = f'danh_sach_ho_kinh_doanh_{date.today()}.xlsx'
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True, download_name=filename)


@export_bp.route('/api/export/pdf')
@login_required
def export_pdf():
    db = get_db()
    with db.cursor() as cur:
        cur.execute("SELECT * FROM ho_kinh_doanh ORDER BY trang_thai, quan_huyen, phuong_xa")
        rows = cur.fetchall()
    return render_template('export_pdf.html', rows=rows, today=date.today())
